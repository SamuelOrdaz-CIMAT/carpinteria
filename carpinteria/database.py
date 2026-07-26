import sqlite3
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except Exception:
    openpyxl = None

from .paths import DB_PATH, SOURCE_XLSX

SCHEMA = """
CREATE TABLE IF NOT EXISTS suppliers (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE,
    contact TEXT DEFAULT '',
    address TEXT DEFAULT '',
    notes TEXT DEFAULT ''
);

CREATE TABLE IF NOT EXISTS materials (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE,
    unit TEXT NOT NULL DEFAULT 'pieza',
    category TEXT DEFAULT '',
    notes TEXT DEFAULT ''
);

CREATE TABLE IF NOT EXISTS material_prices (
    material_id INTEGER NOT NULL,
    supplier_id INTEGER NOT NULL,
    price REAL,
    updated_at TEXT NOT NULL,
    source TEXT DEFAULT '',
    PRIMARY KEY (material_id, supplier_id),
    FOREIGN KEY (material_id) REFERENCES materials(id) ON DELETE CASCADE,
    FOREIGN KEY (supplier_id) REFERENCES suppliers(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS furniture_types (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE,
    description TEXT DEFAULT '',
    labor_cost REAL NOT NULL DEFAULT 0,
    margin_pct REAL NOT NULL DEFAULT 30,
    active INTEGER NOT NULL DEFAULT 1
);

CREATE TABLE IF NOT EXISTS furniture_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    furniture_type_id INTEGER NOT NULL,
    material_id INTEGER NOT NULL,
    quantity REAL NOT NULL DEFAULT 1,
    waste_pct REAL NOT NULL DEFAULT 0,
    preferred_supplier_id INTEGER,
    notes TEXT DEFAULT '',
    FOREIGN KEY (furniture_type_id) REFERENCES furniture_types(id) ON DELETE CASCADE,
    FOREIGN KEY (material_id) REFERENCES materials(id),
    FOREIGN KEY (preferred_supplier_id) REFERENCES suppliers(id)
);

CREATE TABLE IF NOT EXISTS budgets (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    title TEXT NOT NULL,
    customer TEXT DEFAULT '',
    furniture_type_id INTEGER,
    furniture_qty REAL NOT NULL DEFAULT 1,
    labor_cost REAL NOT NULL DEFAULT 0,
    margin_pct REAL NOT NULL DEFAULT 30,
    notes TEXT DEFAULT '',
    created_at TEXT NOT NULL,
    FOREIGN KEY (furniture_type_id) REFERENCES furniture_types(id)
);

CREATE TABLE IF NOT EXISTS budget_lines (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    budget_id INTEGER NOT NULL,
    material_name TEXT NOT NULL,
    unit TEXT NOT NULL,
    supplier_name TEXT DEFAULT '',
    quantity REAL NOT NULL,
    unit_price REAL NOT NULL DEFAULT 0,
    total REAL NOT NULL DEFAULT 0,
    notes TEXT DEFAULT '',
    FOREIGN KEY (budget_id) REFERENCES budgets(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS settings (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL DEFAULT ''
);
"""

DEFAULT_SETTINGS = {
    "workshop_name": "Carpinteria",
    "phone": "",
    "whatsapp": "",
    "address": "",
    "quote_validity": "Vigencia de 7 dias. Precios sujetos a disponibilidad de material.",
    "payment_terms": "Anticipo para iniciar trabajo y liquidacion contra entrega.",
}



def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def initialize() -> None:
    with db() as conn:
        conn.executescript(SCHEMA)
        for key, value in DEFAULT_SETTINGS.items():
            conn.execute("INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)", (key, value))
        material_count = conn.execute("SELECT COUNT(*) FROM materials").fetchone()[0]
    if material_count == 0 and SOURCE_XLSX.exists() and openpyxl:
        import_prices_from_excel(SOURCE_XLSX)


def get_settings(conn: sqlite3.Connection) -> dict:
    rows = conn.execute("SELECT key, value FROM settings").fetchall()
    settings = DEFAULT_SETTINGS.copy()
    settings.update({row["key"]: row["value"] for row in rows})
    return settings

def import_prices_from_excel(path: Path) -> None:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.active
    headers = [sheet.cell(row=8, column=col).value for col in range(1, sheet.max_column + 1)]
    supplier_names = [name for name in headers[2:8] if name]
    today = datetime.now().strftime("%Y-%m-%d")

    with db() as conn:
        supplier_ids = {}
        for name in supplier_names:
            conn.execute("INSERT OR IGNORE INTO suppliers (name) VALUES (?)", (str(name).strip(),))
            supplier_ids[str(name).strip()] = conn.execute(
                "SELECT id FROM suppliers WHERE name = ?", (str(name).strip(),)
            ).fetchone()["id"]

        for row in range(9, 29):
            material_name = sheet.cell(row=row, column=2).value
            if not material_name:
                continue
            note = sheet.cell(row=row, column=10).value or ""
            conn.execute(
                "INSERT OR IGNORE INTO materials (name, notes) VALUES (?, ?)",
                (str(material_name).strip(), str(note).strip()),
            )
            material_id = conn.execute(
                "SELECT id FROM materials WHERE name = ?", (str(material_name).strip(),)
            ).fetchone()["id"]
            if note:
                conn.execute("UPDATE materials SET notes = ? WHERE id = ?", (str(note).strip(), material_id))

            for col, supplier_name in zip(range(3, 9), supplier_names):
                price = sheet.cell(row=row, column=col).value
                if isinstance(price, (int, float)):
                    conn.execute(
                        """
                        INSERT INTO material_prices (material_id, supplier_id, price, updated_at, source)
                        VALUES (?, ?, ?, ?, ?)
                        ON CONFLICT(material_id, supplier_id)
                        DO UPDATE SET price = excluded.price, updated_at = excluded.updated_at
                        """,
                        (material_id, supplier_ids[str(supplier_name).strip()], float(price), today, path.name),
                    )
